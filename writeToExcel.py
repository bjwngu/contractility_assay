import openpyxl as e 
from openpyxl.styles import Alignment, Font 
from openpyxl.drawing.image import Image
import numpy as np 
import statistics 
import math 
import os 
from tkinter import Tk
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename


def extract_data(excel_sheet):
    def rename(results):
    # check if string is float 
        def isfloat(s):
            try:
                float(s)
                return True
            except ValueError:
                return False
        # shorten image id to be just treatment and time id 
        for i in range(0, len(results)):
            split = results[i][0].split("_")
            concat = []
            for s in split: 
                # get treatment id 
                if isfloat(s):
                    concat.append(s)
                # get time id 
                if "min" in s:
                    time_pt = s[:s.index("m")]
                    concat.append(time_pt + "min")
            results[i][0] = "_".join(concat)
    def sort(data_dict):
        for treatment,times in data_dict.items():
            buffer = list(times.keys())
            buffer = list(map(int, buffer))
            sorted = np.argsort(np.array(buffer))
            reordered_times = {}
            for i in range(0,len(buffer)):
                reordered_times[str(buffer[sorted[i]])] = times[str(buffer[sorted[i]])]
            data_dict[treatment] = reordered_times
        return data_dict 
    
    results = [] 
    max_cols = excel_sheet.max_column
    max_rows = excel_sheet.max_row 

    # iterate through each col 
    for i in range(1, max_cols): 
        # find col that corresponds to the start of data entry for one image 
        if (excel_sheet.cell(row=1,column=i).value != None): 
            buffer = [] 
            # get image name 
            name = excel_sheet.cell(row=1,column=i).value
            # get area data 
            buffer.clear() 
            buffer.append(name)
            # add area data from col 
            for row in excel_sheet.iter_rows(min_row=3, max_row=max_rows, min_col=i+1, max_col=i+1, values_only=True):
                if row[0] != None:
                    buffer.append(row[0])  # tuple, get first element
            results.append(buffer)
    
    # rename to simplified format: treatment_time 
    rename(results)

    # reorganize as dict {key:treatment, value:{key:time, value:area}}
    dict_results = {}
    treatments = []
    # -> add treatment
    for set in results: 
        id = set[0].split("_")
        get_treatment = id[0]
        get_timept = id[1].split("m")[0]
        # add treatment if not already
        if get_treatment not in treatments: 
            treatments.append(get_treatment)
            dict_results[get_treatment] = {}
        # add time point and corresponding data 
        dict_results[get_treatment][get_timept] = set[1:]
    
    # reorder so time pts are chronological 
    dict_results = sort(dict_results)
    
    return results, dict_results

def analyze(dict_results):
    percent_change_results = {}  
    # parameter in form dict {key:treatment, value:{key:time, value:area}}
    # calculate (relative to 0min) percent area change 
    # -> iterate through each treatment (key) and times (value=dict)
    for treatment, times in dict_results.items(): 
        percent_change_results[treatment] = {} 
        min0_ref = np.array(times['0'])
        for time, areas in times.items():
            buffer = []
            if time != '0':
                compare = np.array(areas)
                pchange = ((min0_ref-compare)/min0_ref)*100 
                pchange = np.ndarray.round(pchange, 3)
                percent_change = np.ndarray.tolist(pchange)
                new_key = time + "_%Change"
                percent_change_results[treatment][new_key] = percent_change

    return percent_change_results

def write_to_excel(area, analysis, file_name, image_folder_path):

    def align_write(cell, value, sheet):
        write_workbook.active = write_workbook[sheet]
        write_sheet = write_workbook.active
        write_sheet[cell] = value
        write_sheet[cell].alignment = Alignment(horizontal="center")

    def write_by_col(columns, data, tag, threshold):
        treatment_col = 'A'
        cellNum_col = 'D'
        treatment_index = [2]  # row start
        col_index_area = 0
        row_index_area = 1 # row moving 

        for treatment, times in data.items():
            treatment_cell = treatment_col + str(treatment_index[-1])
            align_write(treatment_cell, treatment, "Data")
            
            # write 0, 3, 5, and 10 min data 
            for time, areas in times.items():
                buffer = [time + " " + tag] + areas 
                write_sheet[columns[col_index_area] + str(row_index_area)].font = Font(bold=True)
                for i in range(0, len(buffer)):
                    align_write(columns[col_index_area] + str(row_index_area+i),buffer[i], "Data")
                    if (threshold != 0):
                        if (i==0):
                            align_write(analysis_thresh_col[col_index_area] + str(row_index_area+i),buffer[i], "Data")
                            write_sheet[analysis_thresh_col[col_index_area] + str(row_index_area+i)].font = Font(bold=True)
                        elif (int(buffer[i])>=1):
                            align_write(analysis_thresh_col[col_index_area] + str(row_index_area+i),buffer[i], "Data")

                col_index_area += 1
            col_index_area = 0
            
            # number cells for each treatment 
            num_cells = len(times[list(times.keys())[0]])
            # num_cells = len(times['0'])
            for i in range(0,num_cells):
                align_write(cellNum_col + str(row_index_area+1), i+1, "Data")
                row_index_area += 1
        
            if len(treatment_index) != len(data.keys()):
            # move treatment_index 
                treatment_index.append((num_cells*4) + treatment_index[-1])
                row_index_area = treatment_index[-1]-1
        
        return treatment_index
        
    def add_images():
        # move overlay to root dir
        image_files = [] 
        for f in allfiles:
            split_f = f.split("_")
            if (split_f[0] not in image_files) and (split_f[1]=="0min.png"):
                image_files.append(f)
                src_path = os.path.join(image_folder_path, f)
                dst_path = os.path.join(rootdir, f)
                os.rename(src_path, dst_path)
    
        for i in range(0, len(treatment_index)):
            overlay = Image(image_files[i])
            overlay.height = 350
            overlay.width = 350
            cell = 'B' + str(treatment_index[i])
            overlay.anchor = cell
            write_sheet.add_image(overlay, cell)
        
        return image_files
        
    # TODO: Finish
    def stats():
        def calculate_stats(data):
            # variable init
            results_buffer = [] # results of items 1-5
            above_thresh = [] # for general use 
            for val in data:
                if val>=1:
                    above_thresh.append(val)
            # (1). num of cells in line 
            num_cells = len(data)
            results_buffer.append(num_cells)
            # (2). num of cells that contracted >1%
            num_contract = len(above_thresh)
            results_buffer.append(num_contract)
            # (3). % contract
            p_con = round(num_contract/num_cells*100,3)
            results_buffer.append(p_con)
            # (4). average cont
            try:
                avg = round(statistics.mean(above_thresh),3)
            except: 
                avg = 0
            results_buffer.append(avg)
            # (5). standard error 
            try:
                s_err = round(statistics.stdev(above_thresh)/math.sqrt(num_contract), 3)
            except:
                s_err = 0
            results_buffer.append(s_err)

            return results_buffer
        write_workbook.create_sheet(index = 1, title="Results")
        write_workbook.active = write_workbook["Results"]
        write_sheet = write_workbook.active
        columns = ["Cell Line", "# Cells", "#Cont>1%: 3|5|10 min", "%Cont: 3|5|10 min", "Avg%Cont: 3|5|10 min", "StdErr: 3|5|10 min"]
        col_loc = ['B', 'C', 'D', 'G', 'J', 'M']
        
        # calcualte values 
        last_treatment = int(list(analysis.keys())[-1].split(".")[0])
        lines = list(range(1,last_treatment+1)) # ex: [1,2,3,4]
        stats_dict = {} # final result dict 
        for line in lines:
            stats_dict[line] = {'3_%Change':[], '5_%Change':[], '10_%Change':[]} # list of statistical values 
            buffer_dict = {'3_%Change':[], '5_%Change':[], '10_%Change':[]} # unprocessed area values for each time 
            for treatment, times in analysis.items(): # (ex): treatment = 1.1, times = 3_%Change, etc. 
                if int(treatment.split(".")[0]) == line:
                    for time, values in times.items():
                        buffer_dict[time] += values
            for times in stats_dict[line]:
                stats_dict[line][times] += calculate_stats(buffer_dict[times])
        
        # write to new sheet 
        # -> init col names 
        for i in range(0, len(columns)):
            start_row = '2'
            write_sheet[col_loc[i]+start_row] = columns[i]
            write_sheet[col_loc[i]+start_row].alignment = Alignment(horizontal="center")
            write_sheet[col_loc[i]+start_row].font = Font(bold=True)
            write_sheet.column_dimensions[col_loc[i]].width = 10                
        
        # -> merge at DEF, GHI, JKL, and MNO
        write_sheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=6)
        write_sheet.merge_cells(start_row=2, start_column=7, end_row=2, end_column=9)
        write_sheet.merge_cells(start_row=2, start_column=10, end_row=2, end_column=12)
        write_sheet.merge_cells(start_row=2, start_column=13, end_row=2, end_column=15)

        # -> write values 
        for line in lines:
            # write line and num_cells
            align_write('B'+str(line+2), line, "Results")
            align_write('C'+str(line+2), stats_dict[line]['3_%Change'][0], "Results")
            min_3_val = ['D', 'G', 'J', 'M']
            min_5_val = ['E', 'H', 'K', 'N']
            min_10_val = ['F', 'I', 'L', 'O']
            for time, value in stats_dict[line].items():
                if time == '3_%Change':
                    buffer = min_3_val
                elif time == '5_%Change':
                    buffer = min_5_val
                else:
                    buffer = min_10_val
                for col in range(0, len(buffer)):
                    align_write(buffer[col]+str(line+2), value[col+1], "Results")
    
    # init workbook for writing 
    write_workbook = e.Workbook() 
    write_sheet = write_workbook.active 
    write_sheet.title = "Data"

    #format cells 
    areas_col = ['E','F','H','J']
    analysis_col = ['G','I','K']
    analysis_thresh_col = ['M','N','O']
    write_sheet.column_dimensions['A'].width = 10
    write_sheet.column_dimensions['B'].width = 50 # to fit image overlay 
    for i in range(0, len(analysis_col)):
        write_sheet.column_dimensions[analysis_col[i]].width = 15
        write_sheet.column_dimensions[analysis_thresh_col[i]].width = 15
    write_sheet["B1"].alignment = Alignment(horizontal="center")
    write_sheet["A1"] = "Treatment"
    write_sheet["A1"].font = Font(bold=True)
    write_sheet["B1"] = "Image + ROI"
    write_sheet["B1"].font = Font(bold=True)

    # get image files 
    rootdir = os.getcwd()
    allfiles = os.listdir(image_folder_path)
    
    # function calls for writing 
    treatment_index = write_by_col(areas_col, area, " min", 0)
    treatment_index = write_by_col(analysis_col, analysis, "", 1)
    image_files = add_images() 
    stats() 

    # end of function
    try:
        write_workbook.save(filename=file_name)
        # return images to its own folder in root 
        for f in image_files:
            src_path = os.path.join(rootdir, f)
            dst_path = os.path.join(image_folder_path, f)
            os.rename(src_path, dst_path)
    except:
        print("did not save")
        # return images to its own folder in root 
        for f in image_files:
            src_path = os.path.join(rootdir, f)
            dst_path = os.path.join(image_folder_path, f)
            os.rename(src_path, dst_path)
      
def main():
    print("Cell Segmentation Data Analysis")
    read_file_name = askopenfilename(title='Select read file (.xlsx)')
    image_folder_path = askdirectory(title='Select Image Folder') 
    write_file_name = input("Enter write file name (.xlsx): ")
  
    # read data init. 
    read_workbook = e.load_workbook(read_file_name) 
    read_sheet = read_workbook.active 
    data, dict_data = extract_data(read_sheet)
    dict_data_analyzed = analyze(dict_data)
    write_to_excel(dict_data, dict_data_analyzed, write_file_name, image_folder_path)
   
    
if __name__ == '__main__':
    main()











